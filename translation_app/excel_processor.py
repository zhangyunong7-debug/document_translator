"""
Excel file processor for translation
"""
import re
from typing import List, Tuple, Callable, Optional
from openpyxl import load_workbook
from openpyxl.cell import Cell


class ExcelProcessor:
    """Processor for Excel files (.xlsx, .xls)"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.translated_values = {}  # Cache for translated values

    def load(self) -> bool:
        """Load the Excel workbook"""
        try:
            self.workbook = load_workbook(self.file_path)
            return True
        except Exception as e:
            raise Exception(f"Failed to load Excel file: {str(e)}")

    def extract_all_cells(self) -> List[Tuple[Cell, any]]:
        """
        Extract all cells with content from the workbook
        Returns list of tuples: (cell, original_value)
        """
        cells = []

        if not self.workbook:
            return cells

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and self._is_text_content(cell):
                        cells.append((cell, cell.value))

        return cells

    def _is_text_content(self, cell: Cell) -> bool:
        """Check if cell contains text content worth translating"""
        value = cell.value
        if value is None:
            return False

        # Check if it's a string or can be converted to string
        value_str = str(value).strip()

        # Skip if it looks like a formula
        if value_str.startswith('=') or value_str.startswith('{='):
            return False

        # Skip pure numbers
        if value_str.replace('.', '').replace('-', '').replace(',', '').isdigit():
            return False

        # Skip dates and other special values
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return False

        return len(value_str) > 0

    def translate_content(self, translator, source_lang: str, target_lang: str,
                          progress_callback: Optional[Callable] = None) -> bool:
        """
        Translate all cell content in the workbook
        """
        if not self.workbook:
            raise Exception("Workbook not loaded")

        cells = self.extract_all_cells()
        total = len(cells)
        self.translated_values = {}

        if total == 0:
            return True

        for index, (cell, original_value) in enumerate(cells):
            try:
                original_str = str(original_value)
                translated = translator.translate(original_str, source_lang, target_lang)
                cell.value = translated
                self.translated_values[(cell.coordinate, original_str)] = translated
            except Exception as e:
                print(f"Translation error for cell {cell.coordinate}: {e}")

            if progress_callback:
                progress_callback(index + 1, total)

        return True

    def save(self, output_path: str) -> bool:
        """Save the translated workbook"""
        try:
            self.workbook.save(output_path)
            return True
        except Exception as e:
            raise Exception(f"Failed to save Excel file: {str(e)}")

    def get_sheet_count(self) -> int:
        """Get number of sheets in workbook"""
        if self.workbook:
            return len(self.workbook.sheetnames)
        return 0

    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names"""
        if self.workbook:
            return self.workbook.sheetnames.copy()
        return []
